"""Export `RagEvalRun`s to Excel or JSON for hand-off to stakeholders.

The Excel workbook has one sheet per system_name plus a comparison sheet
when two or more systems are present. JSON export is a single document
keyed by system_name — useful for ad-hoc analysis in notebooks.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from multihop_eval.rag_eval.models import RagEvalRun

_HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_kv_block(ws, start_row: int, title: str, data: dict[str, float]) -> int:
    """Write a 'title' header + `metric | value` rows; return next free row."""
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12)
    start_row += 1
    ws.cell(row=start_row, column=1, value="metric").font = _HEADER_FONT
    ws.cell(row=start_row, column=2, value="value").font = _HEADER_FONT
    ws.cell(row=start_row, column=1).fill = _HEADER_FILL
    ws.cell(row=start_row, column=2).fill = _HEADER_FILL
    ws.cell(row=start_row, column=1).alignment = _HEADER_ALIGN
    ws.cell(row=start_row, column=2).alignment = _HEADER_ALIGN
    row = start_row + 1
    for metric, value in sorted(data.items()):
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2, value=float(value) if value is not None else None)
        row += 1
    return row + 1


def _write_per_query_block(ws, start_row: int, rows: list[dict[str, Any]]) -> int:
    """Write the per-query drill-down with one header row + one row per response."""
    if not rows:
        return start_row
    columns = sorted({k for row in rows for k in row}, key=lambda c: (c != "qa_pair_key", c))
    ws.cell(row=start_row, column=1, value="Per-query drill-down").font = Font(bold=True, size=12)
    start_row += 1
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
    row = start_row + 1
    for record in rows:
        for col_idx, name in enumerate(columns, start=1):
            value = record.get(name)
            # Coerce booleans to 0/1 so they sort correctly in Excel filters.
            if isinstance(value, bool):
                value = int(value)
            ws.cell(row=row, column=col_idx, value=value)
        row += 1
    return row + 1


def export_rag_eval_to_excel(
    runs: list[RagEvalRun], output_path: str | Path
) -> Path:
    """Write every run + an aggregate comparison sheet to an .xlsx file.

    Args:
        runs: One or more `RagEvalRun`s. With zero runs we still produce an
            empty workbook so callers don't need to guard against that case.
        output_path: Destination `.xlsx` path; parent dirs are created.

    Returns:
        The resolved output `Path`.
    """
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # `Workbook()` starts with one default sheet; we rename it as the cover
    # so the workbook never has an awkward 'Sheet' tab dangling.
    cover = wb.active
    cover.title = "Summary"
    cover.cell(row=1, column=1, value="System").font = _HEADER_FONT
    cover.cell(row=1, column=1).fill = _HEADER_FILL
    cover.cell(row=1, column=2, value="N responses").font = _HEADER_FONT
    cover.cell(row=1, column=2).fill = _HEADER_FILL
    cover.cell(row=1, column=3, value="N matched").font = _HEADER_FONT
    cover.cell(row=1, column=3).fill = _HEADER_FILL
    for row_idx, run in enumerate(runs, start=2):
        cover.cell(row=row_idx, column=1, value=run.system_name)
        cover.cell(row=row_idx, column=2, value=run.n_responses)
        cover.cell(row=row_idx, column=3, value=run.n_matched_goldens)

    for run in runs:
        sheet_name = run.system_name[:31] or "system"
        ws = wb.create_sheet(title=sheet_name)
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 18
        next_row = 1
        next_row = _write_kv_block(ws, next_row, "Retrieval metrics", run.metrics.retrieval)
        next_row = _write_kv_block(ws, next_row, "Generation metrics", run.metrics.generation)
        _write_per_query_block(ws, next_row, run.metrics.per_query)

    wb.save(out)
    return out


def export_rag_eval_to_json(
    runs: list[RagEvalRun], output_path: str | Path
) -> Path:
    """Dump runs to a single JSON file keyed by `system_name`.

    The schema is the same as `RagEvalRun.model_dump()` with `datetime`
    fields ISO-formatted so the file is portable to notebooks.
    """
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        run.system_name: json.loads(run.model_dump_json()) for run in runs
    }
    out.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    return out
