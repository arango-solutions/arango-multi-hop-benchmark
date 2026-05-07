"""Write a list of `AcceptedQA` rows to a styled Excel workbook."""

from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from multihop_eval.models import AcceptedQA

log = logging.getLogger(__name__)

HOP_COLOURS: dict[int, str] = {2: "DDEBF7", 3: "E2EFDA", 4: "FFF2CC", 5: "FCE4D6"}
DEFAULT_OUT_OF_RANGE_COLOUR = "FFFFFF"


def export_to_excel(
    rows: list[AcceptedQA],
    output_path: str | Path,
    *,
    sheet_title: str = "Multi-Hop Eval",
    rubric_field_names: list[str] | None = None,
) -> Path:
    """Write `rows` to an .xlsx file at `output_path` and return the path.

    The header row is dark navy with white bold text; data rows are coloured by
    hop count (2-hop → blue, 3-hop → green, ...). When `rubric_field_names` is
    supplied, one column per field is appended with the per-field score; an
    additional 'rubric_weighted_score' column is always appended when at least
    one row carries a weighted score.
    """
    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # Excel sheet title hard-cap

    base_headers = [
        "cluster_id",
        "partition_id",
        "hop_count",
        "persona",
        "reasoning_chain",
        "question",
        "golden_answer",
        "proof",
    ]
    rubric_field_names = rubric_field_names or _collect_rubric_fields(rows)
    rubric_columns = [f"rubric.{f}" for f in rubric_field_names]
    has_weighted = any(r.rubric_weighted_score is not None for r in rows)
    headers = list(base_headers)
    headers.extend(rubric_columns)
    if has_weighted:
        headers.append("rubric_weighted_score")

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for ri, row in enumerate(rows, start=2):
        rd = row.to_row_dict()
        fill = PatternFill(
            "solid",
            fgColor=HOP_COLOURS.get(row.hop_count, DEFAULT_OUT_OF_RANGE_COLOUR),
        )
        base_values: list[Any] = [
            rd["cluster_id"],
            rd["partition_id"],
            rd["hop_count"],
            rd["persona"],
            rd["reasoning_chain"],
            rd["question"],
            rd["answer"],
            rd["proof"],
        ]
        for f in rubric_field_names:
            entry = row.rubric_scores.get(f)
            if entry is None:
                base_values.append("")
            else:
                base_values.append(
                    f"{entry.score} — {entry.justification}"
                )
        if has_weighted:
            base_values.append(
                f"{row.rubric_weighted_score:.3f}" if row.rubric_weighted_score is not None else ""
            )

        for ci, value in enumerate(base_values, start=1):
            cell = ws.cell(row=ri, column=ci, value=_safe_cell_value(value))
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    base_widths = [30, 12, 10, 18, 40, 55, 55, 60]
    rubric_widths = [25] * len(rubric_columns)
    weighted_width = [18] if has_weighted else []
    for ci, w in enumerate(base_widths + rubric_widths + weighted_width, start=1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w

    wb.save(out_path)
    log.info("Saved %d rows -> %s", len(rows), out_path)
    return out_path


def _collect_rubric_fields(rows: list[AcceptedQA]) -> list[str]:
    seen: list[str] = []
    for r in rows:
        for k in r.rubric_scores:
            if k not in seen:
                seen.append(k)
    return seen


def _safe_cell_value(value: Any) -> Any:
    """Excel can't store dicts/lists; fall back to JSON for those."""
    if isinstance(value, (dict, list)):
        return json.dumps(value, indent=2)
    return value
