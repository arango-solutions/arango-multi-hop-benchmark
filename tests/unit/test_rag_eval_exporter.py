"""Tests for `multihop_eval.exporters.rag_eval_exporter`."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from multihop_eval.exporters.rag_eval_exporter import (
    export_rag_eval_to_excel,
    export_rag_eval_to_json,
)
from multihop_eval.rag_eval.models import RagEvalRun, RagMetricBundle


def _run(system: str) -> RagEvalRun:
    return RagEvalRun(
        system_name=system,
        n_responses=2,
        n_matched_goldens=2,
        metrics=RagMetricBundle(
            retrieval={"precision@1": 0.5, "mrr": 0.75},
            generation={"rouge_l_f1": 0.4, "groundedness": 0.8},
            per_query=[
                {"qa_pair_key": "q1", "precision@1": 1.0, "groundedness": 1.0},
                {"qa_pair_key": "q2", "precision@1": 0.0, "groundedness": 0.6},
            ],
        ),
    )


def test_excel_export_creates_summary_and_per_system_sheets(tmp_path: Path):
    runs = [_run("rag_a"), _run("rag_b")]
    out = export_rag_eval_to_excel(runs, tmp_path / "rag_eval.xlsx")
    assert out.exists()

    wb = load_workbook(out)
    assert "Summary" in wb.sheetnames
    assert "rag_a" in wb.sheetnames
    assert "rag_b" in wb.sheetnames

    summary = wb["Summary"]
    # Row 1 = headers, rows 2..3 = systems.
    assert summary.cell(row=1, column=1).value == "System"
    systems_listed = {summary.cell(row=r, column=1).value for r in (2, 3)}
    assert systems_listed == {"rag_a", "rag_b"}


def test_excel_export_sheet_holds_metric_blocks(tmp_path: Path):
    runs = [_run("solo")]
    out = export_rag_eval_to_excel(runs, tmp_path / "x.xlsx")
    wb = load_workbook(out)
    ws = wb["solo"]
    # First block heading: "Retrieval metrics".
    assert ws.cell(row=1, column=1).value == "Retrieval metrics"
    # Find the per-query header row by walking down column A.
    header_rows = [
        row[0].value
        for row in ws.iter_rows(min_col=1, max_col=1)
        if row[0].value
        in {"Retrieval metrics", "Generation metrics", "Per-query drill-down"}
    ]
    assert header_rows == [
        "Retrieval metrics",
        "Generation metrics",
        "Per-query drill-down",
    ]


def test_excel_export_handles_zero_runs(tmp_path: Path):
    out = export_rag_eval_to_excel([], tmp_path / "empty.xlsx")
    wb = load_workbook(out)
    assert wb.sheetnames == ["Summary"]


def test_excel_export_truncates_long_system_names(tmp_path: Path):
    long_name = "rag_with_a_really_long_system_name_exceeding_excel_limit_of_thirty_one"
    out = export_rag_eval_to_excel([_run(long_name)], tmp_path / "long.xlsx")
    wb = load_workbook(out)
    # Excel limits sheet names to 31 chars — exporter must accommodate.
    assert any(len(name) <= 31 for name in wb.sheetnames)


def test_json_export_round_trip(tmp_path: Path):
    runs = [_run("rag_a"), _run("rag_b")]
    out = export_rag_eval_to_json(runs, tmp_path / "rag_eval.json")
    payload = json.loads(out.read_text(encoding="utf-8"))
    assert set(payload) == {"rag_a", "rag_b"}
    assert payload["rag_a"]["metrics"]["retrieval"]["mrr"] == 0.75
    assert payload["rag_a"]["n_responses"] == 2


def test_json_export_iso_dates(tmp_path: Path):
    out = export_rag_eval_to_json([_run("rag_a")], tmp_path / "rag_eval.json")
    payload = json.loads(out.read_text(encoding="utf-8"))
    # ISO-8601 always contains a 'T' separator between date and time.
    assert "T" in payload["rag_a"]["started_at"]


def test_excel_export_creates_parent_dirs(tmp_path: Path):
    nested = tmp_path / "deeply" / "nested" / "out.xlsx"
    out = export_rag_eval_to_excel([_run("rag_a")], nested)
    assert out.exists()
