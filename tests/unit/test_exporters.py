"""Tests for `multihop_eval.exporters`."""

from __future__ import annotations

import json
from datetime import UTC, datetime

from openpyxl import load_workbook

from multihop_eval.exporters import export_to_excel, export_to_json
from multihop_eval.exporters.json_exporter import export_run_to_json
from multihop_eval.models import (
    AcceptedQA,
    ProofPoint,
    RejectedQA,
    RejectionReason,
    RubricScore,
    RunResult,
)


def _accepted(
    *,
    hops: int = 2,
    cluster: str = "dom/cluster_a",
    persona: str = "hr_manager",
    rubric_scores: dict | None = None,
    weighted: float | None = None,
) -> AcceptedQA:
    return AcceptedQA(
        cluster_id=cluster,
        partition_id="part_0",
        hop_count=hops,
        persona=persona,
        reasoning_chain="A->B",
        question="Test question?",
        answer="Plain prose answer.",
        proof_list=[
            ProofPoint(point="fact 1", source_id="src/aaa"),
            ProofPoint(point="fact 2", source_id="src/bbb"),
        ],
        rubric_scores=rubric_scores or {},
        rubric_weighted_score=weighted,
    )


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------


def test_excel_export_writes_headers_and_rows(tmp_path):
    rows = [_accepted(hops=2), _accepted(hops=3)]
    out = export_to_excel(rows, tmp_path / "out.xlsx")
    assert out.exists()
    wb = load_workbook(out)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    for expected in ["cluster_id", "hop_count", "question", "golden_answer", "proof"]:
        assert expected in headers
    # Two data rows in addition to the header.
    assert ws.max_row == 3


def test_excel_export_colours_rows_by_hop_count(tmp_path):
    rows = [_accepted(hops=2), _accepted(hops=3)]
    out = export_to_excel(rows, tmp_path / "out.xlsx")
    wb = load_workbook(out)
    ws = wb.active
    # Header row uses 1F4E79 background.
    header_fill = ws.cell(row=1, column=1).fill.fgColor.rgb
    assert "1F4E79" in str(header_fill)
    fill_2hop = ws.cell(row=2, column=1).fill.fgColor.rgb
    fill_3hop = ws.cell(row=3, column=1).fill.fgColor.rgb
    assert fill_2hop != fill_3hop


def test_excel_export_appends_rubric_columns_when_present(tmp_path):
    rows = [
        _accepted(
            rubric_scores={
                "factuality": RubricScore(5, "good"),
                "conciseness": RubricScore(4, "ok"),
            },
            weighted=0.92,
        ),
    ]
    out = export_to_excel(rows, tmp_path / "out.xlsx")
    wb = load_workbook(out)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert "rubric.factuality" in headers
    assert "rubric.conciseness" in headers
    assert "rubric_weighted_score" in headers
    fact_col = headers.index("rubric.factuality") + 1
    assert "5" in str(ws.cell(row=2, column=fact_col).value)
    assert "good" in str(ws.cell(row=2, column=fact_col).value)


def test_excel_export_handles_empty_rows_list(tmp_path):
    out = export_to_excel([], tmp_path / "empty.xlsx")
    wb = load_workbook(out)
    ws = wb.active
    assert ws.max_row == 1  # header only


# ---------------------------------------------------------------------------
# JSON
# ---------------------------------------------------------------------------


def test_json_export_round_trips_rows(tmp_path):
    rows = [_accepted(rubric_scores={"factuality": RubricScore(5, "ok")}, weighted=0.8)]
    out = export_to_json(rows, tmp_path / "out.json")
    payload = json.loads(out.read_text(encoding="utf-8"))
    assert isinstance(payload, list)
    assert payload[0]["question"] == "Test question?"
    assert payload[0]["rubric_scores"]["factuality"]["score"] == 5
    assert payload[0]["rubric_weighted_score"] == 0.8


def test_export_run_to_json_includes_rejected_and_timing(tmp_path):
    started = datetime(2026, 1, 1, 0, 0, 0, tzinfo=UTC)
    finished = datetime(2026, 1, 1, 0, 0, 30, tzinfo=UTC)
    result = RunResult(
        accepted=[_accepted()],
        rejected=[
            RejectedQA(
                cluster_id="dom/cluster_a",
                persona="hr_manager",
                seed_doc_id="src/x",
                reason=RejectionReason.MULTIHOP_BELOW_FLOOR,
            )
        ],
        cluster_targets={"dom/cluster_a": 5},
        cluster_achieved={"dom/cluster_a": 1},
        started_at=started,
        finished_at=finished,
    )
    out = export_run_to_json(result, tmp_path / "run.json")
    data = json.loads(out.read_text())
    assert data["duration_s"] == 30.0
    assert data["accepted"][0]["question"] == "Test question?"
    assert data["rejected"][0]["reason"] == RejectionReason.MULTIHOP_BELOW_FLOOR.value
